from openpyxl import load_workbook
from imaging_type import *


# path = 'pandas_to_excel.xlsx'


def write_to_xlsx(path, data, row, column, sheet_index):
    wb = load_workbook(path)
    sheets = wb.sheetnames
    dr_note = wb[sheets[sheet_index]]
    # Then update as you want it
    dr_note.cell(row=row, column=column).value = data  # This will change the cell(2,4) to 4
    wb.save("NEW EXCEL SAMPLE.xlsx")


def body_area_line(filename, area):
    # open the sample text representing the imaging doc
    doc = open(filename)

    # read the content of the file opened
    content = doc.readlines()

    # pre-proccesing, could be in another function
    fixed_content = [x.replace('\n', '') for x in content if x != '\n']

    start_of_content = fixed_content[0:15]
    index = 0
    for line in start_of_content:
        index += 1
        # makes sure it's not part of another word
        if area + ' ' in line or ' ' + area in line:
            return index
    return float('inf')


def body_area(filename, body_areas, imaging_type):
    if imaging_type == "PET CT":
        return 'whole body'
    line_list = [float('inf')] * len(body_areas)
    index = -1
    for area in body_areas:
        index += 1
        line_list[index] = body_area_line(filename, area)
    return body_areas[line_list.index(min(line_list))]


def specific_sites_pages(doc_pages, poss_sites, keys):
    sites = []
    for page in doc_pages:
        sites += specific_sites(page, poss_sites, keys)
    return sites

def specific_sites(doc, poss_sites, keys):
    sites = []
    # pre-proccesing, could be in another function
    poss_sites = [s.lower() for s in poss_sites]
    # open the sample text representing the imaging doc
    doc = open(doc)

    # read the content of the file opened
    content = doc.readlines()

    # pre-proccesing, could be in another function
    fixed_content = [x.replace('\n', '') for x in content if x != '\n']

    index = 0
    for line in fixed_content:
        index += 1
        if any(k in line for k in keys):
            break

    sites_content = fixed_content[index:]
    # pre-proccesing, could be in another function
    sites_content = [s.lower() for s in sites_content]
    for site in poss_sites:
        for line in sites_content:
            if site + '/' in line or site + ':' in line or site + ' ' in line or site + ',' in line:
                sites.append(site)
    return sites


# could be improve using dictionary with values as other ways to write
imaging_types = ['XR', 'MRI', 'PET CT', 'CT', 'DOTATATE']
body_areas = ['brain to thigh', 'brain to chest', 'chest and abdomen', 'abdomen and pelvis', 'brain', 'chest',
              'abdomen', 'bone', 'whole body']
sites_keys = ['FINDINGS', 'REPORT', 'IMPRESSION']
poss_sites = ['Chest lymph nodes', 'Adenopathy', 'Lower Neck', 'Thyroid gland', 'Trachea', 'Supracalvicular',
              'Pulmonary Arteries', 'Pulmonary Parenchyma', 'Airways', 'Pleural Space', 'Heart', 'Aorta',
              'Pericardium', 'Mediastinum', 'Hila', 'Thoracic Vessels', 'Osseous Structures', 'Chest Wall',
              'Costosternal', 'Mammary node', 'Breast', 'retrocaval', 'Upper Abdomen', 'Right lobe',
              'Lung', 'Thyroid', 'Lungs', 'airway and pleura', 'Mediastinum and hila', 'Heart and great vessels',
              'Breast', 'Pericardial or pleural', 'Extra-axial spaces', 'Intracranial hemorrhage',
              'Ventircular system', 'Basal cisterns', 'Cerebral parenchyma', 'Midline shift', 'Cerebellum',
              'Brainstem', 'Calavarium', 'Vascular system', 'Paranalsal sinuses and mastoid air cells',
              'Visualized orbits', 'Visualized upper cervical spine', 'Sella', 'Skull base', 'Marrow',
              'Frontal', 'Hypermetabolic foci', 'Liver', 'Gallbladder', 'Bile ducts', 'Kidney', 'Adrenal',
              'Spleen', 'Pancreas', 'Great abdominal vessels', 'Stomach', 'Bowel', 'Appendix', 'Pelvis',
              'Mesentry/Peritoneum', 'Intestine', 'Vasulature', 'Bone windows', 'Gastrointestinal trac',
              'Groin', 'Colon', 'Lung bases', 'Retroperitoneum', 'Rectum', 'Spine', 'Bones and soft tissue',
              'Vertebrae', 'Skeleton', 'Head and neck', 'Abdomen and pelvis', 'Thorax', 'Musculoskeletal',
              'Thoracic inlet', 'Breast and axilla', 'Kidneys, ureters', 'Gastrointestinal tract', 'Mesentery',
              'Peritoneum', 'Retroperitoneum', 'Vasculature']

# area1 = body_area('doc1.txt', body_areas, imaging_type('doc1.txt', imaging_types))
# area2 = body_area('doc2.txt', body_areas, imaging_type('doc2.txt', imaging_types))
# area3 = body_area('doc3.txt', body_areas, imaging_type('doc3.txt', imaging_types))
# print(area1)
# print(area2)
# print(area3)

doc1_pages = ['doc1.txt']
doc2_pages = ['doc2.txt']
doc3_pages = ['doc3_1.txt', 'doc3_2.txt', 'doc3_3.txt']
sites1 = specific_sites_pages(doc1_pages, poss_sites, sites_keys)
# sites2 = specific_sites_pages(doc2_pages, poss_sites, sites_keys)
sites3 = specific_sites_pages(doc3_pages, poss_sites, sites_keys)
print(sites1)
# print(sites2)
print(sites3)

# write_to_xlsx('ac example.xlsx', area1, 19, 19)
# write_to_xlsx('ac example.xlsx', area2, 20, 19)
# write_to_xlsx('ac example.xlsx', area3, 21, 19)
