"""
Python script for recognizing the imaging type of imaging medical documents
"""

__name__ = "imaging_type"
__author__ = "Gal Peled"

from openpyxl import load_workbook

# path = 'pandas_to_excel.xlsx'

# ignore this function for now, excel work has been transferred to Enav
def write_to_xlsx(path, typ, row, column):
    # file = 'input.xlsx'
    # new_row = ['data1', 'data2', 'data3', 'data4']
    #
    # wb = openpyxl.load_workbook(filename=filename)
    # ws = wb['Sheet1']  # Older method was  .get_sheet_by_name('Sheet1')
    # row = ws.get_highest_row() + 1
    #
    # for col, entry in enumerate(new_row, start=1):
    #     ws.cell(row=row, column=col, value=entry)
    #
    # wb.save(file)

    # wb = load_workbook(filename=filename)
    # sheet_ranges = wb['range names']
    # print(sheet_ranges['D18'].value)

    # path = 'pandas_to_excel.xlsx'
    #
    # with pd.ExcelWriter(path) as writer:
    #     writer.book = openpyxl.load_workbook(path)
    #     df.to_excel(writer, sheet_name='new_sheet1')
    #     df2.to_excel(writer, sheet_name='new_sheet2')

    wb = load_workbook(path)
    sheets = wb.sheetnames
    dr_note = wb[sheets[4]]
    # Then update as you want it
    dr_note.cell(row=row, column=column).value = typ  # This will change the cell(2,4) to 4
    wb.save("NEW EXCEL SAMPLE.xlsx")


def imaging_type_line(filename, typ):
    """
    Finds and returns the line index where the imaging type typ appears in the imaging document txt file.
    Does it by iterating over the lines in the file and finding where typ is.

    Parameters
    ----------
    filename : str
        A string representing the name of the txt file generated by OCR process on an imaging document
    typ : str
        A string representing one imaging type

    Returns
    -------
    int
        Index of typ in the txt file

    """
    # open the sample text representing the imaging doc
    doc = open(filename)

    # read the content of the file opened
    content = doc.readlines()

    # pre-processing for deleting empty lines generated by OCR, could be in another function
    fixed_content = [x.replace('\n', '') for x in content if x != '\n']

    # take only the first 15 lines of the file to recognize the headline where the imaging type should be
    start_of_content = fixed_content[0:15]
    index = 0
    for line in start_of_content:
        index += 1
        # makes sure it's not part of another word
        if typ+' ' in line or ' '+typ in line:
            return index
    # if not found typ in the file, return infinity index because later we want the minimal index
    return float('inf')


def imaging_type(filename, poss_imaging_types):
    """
    Finds and returns type of the imaging document.
    Does it by finding the minimal index of a possible imaging type in the imaging document txt file.

    Parameters
    ----------
    filename : str
        A string representing the name of the txt file generated by OCR proccess on an imaging document
    poss_imaging_types : list
        A list representing a closed set of possible imaging types

    Returns
    -------
    str
        The most probably imaging type of the document

    """
    # initialize a list of infinity values, of the same length as the length of poss_imaging_types
    # each value in the list will finally be the index of the line in the file, or infinity if not found
    line_list = [float('inf')] * len(poss_imaging_types);
    index = -1
    for typ in poss_imaging_types:
        index += 1
        line_list[index] = imaging_type_line(filename, typ)
    # return the imaging type with the minimal line index of a possible type, which is most probably the real imaging type
    return poss_imaging_types[line_list.index(min(line_list))]



# could be improve using dictionary with values as other ways to write
poss_imaging_types = ['XR', 'MRI', 'PET CT', 'CT', 'DOTATATE']
# type1 = imaging_type('doc1.txt', poss_imaging_types)
# type2 = imaging_type('doc2.txt', poss_imaging_types)
# type3 = imaging_type('doc3.txt', poss_imaging_types)
# print(type1)
# print(type2)
# print(type3)

# write_to_xlsx('ac example.xlsx', type1, 19, 19)
# write_to_xlsx('ac example.xlsx', type2, 20, 19)
# write_to_xlsx('ac example.xlsx', type3, 21, 19)

